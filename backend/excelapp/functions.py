import openpyxl
import pandas as pd
from glob import glob

filepath = 'sample.xlsx'

# pandasのdataframe形式で読み取ってtable表示
# jQueryでタブをつまんで列順を並べ変える/またはとってくる列の順番をチェックを入れブラウザで選択する　
#0,1,2,3,4,にカラム名1,2,3,4が対応
# a,b,c,dには列番号を入力してください 表示したい順番に列が並びます
def arrange(filepath, a, b, c, d):
  _df = pd.read_excel(filepath)
  df = _df.iloc[0: , [a+1, b+1, c+1, d+1]]
  return df


# 全ファイルで同じ形式で結合
  df = pd.DataFrame()
  for filepath in filepaths:
    _df = arrange(filepath)
    df = pd.concat([df, _df])
  df.to_excel('arranged_data.xlsx', index=False)


#A2以降にnameがある
def sheetbyname(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.worksheets[0]
    namelist = set()
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        name = row[0].value
        if isinstance(name, str) and name.strip():
            namelist.add(name[:31])  # シート名の最大長は31文字
    for name in namelist:
        wb.create_sheet(title=name)
    return wb


#A2以降にnameがある
def copyrowdata(file_path):
  wb = openpyxl.load_workbook(file_path)
  ws = wb.active
  for row_no in range(2, ws.max_row + 1):
    name = ws.cell(row=row_no, column=1).value
    if name and name in wb.sheetnames:
      target_sheet = wb[name]
      new_row = target_sheet.max_row + 1
      for col_no, cell in enumerate(ws[row_no], start=1):
        target_sheet.cell(row=new_row, column=col_no, value=cell.value)
  return wb


# from openpyxl.worksheet.worksheet import Worksheet
# PAPERSIZE_LETTER = '1'
# PAPERSIZE_LETTER_SMALL = '2'
# PAPERSIZE_TABLOID = '3'
# PAPERSIZE_LEDGER = '4'
# PAPERSIZE_LEGAL = '5'
# PAPERSIZE_STATEMENT = '6'
# PAPERSIZE_EXECUTIVE = '7'
# PAPERSIZE_A3 = '8'
# PAPERSIZE_A4 = '9'
# PAPERSIZE_A4_SMALL = '10'
# PAPERSIZE_A5 = '11'
# B5の場合
# ws.page_setup.paperSize.customPaperSize = (182, 257)

# printinfo = PrintInfo()
# # クラスからとってきた値を代入する関数にする, 例えば,papersize=10なら10を代入する

# def excel_print(filepath, print_info): #　インスタンスごと渡す
#   wb = openpyxl.load_workbook(filepath)
#   for ws in wb.worksheets:
#       ws.page_setup.paperSize =  printinfo.papersize #数値でok
#       ws.page_setup.orientation =  "f{printinfo.orientation}" #string
#       ws.page_setup.fitToWidth = printinfo.fitToWidth #横なら1自動なら0
#       ws.page_setup.fitToHeight = printinfo.fidToHeight #縦1自動なら0
#       ws.sheet_properties.pageSetUpPr.fitToPage = True
#   return wb

# そもそもの定義が違う。関数の引数でいきなりインスタンス変数を渡すことはできないので
# 関数にはシンプルな変数だけを引数に渡すようにする
# 特に、インスタンス変数と名前を同期させるのが良い


# 印刷揃える機能
def excel_print(filepath, papersize, orientation, fitToWidth, fitToHeight):
  wb = openpyxl.load_workbook(filepath)
  for ws in wb.worksheets:
      ws.page_setup.paperSize =  papersize #数値でok
      ws.page_setup.orientation =  "f{orientation}" #string
      ws.page_setup.fitToWidth = fitToWidth #横なら1自動なら0
      ws.page_setup.fitToHeight = fitToHeight #縦1自動なら0
      ws.sheet_properties.pageSetUpPr.fitToPage = True
  return wb