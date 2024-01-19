import openpyxl
import pandas as pd
from glob import glob

from app import PrintInfo

filepath = 'sample.xlsx'


# a,b,c,dには列番号を入力してください 表示したい順番に列が並びます
def arrange(filepath, a, b, c, d):
  _df = pd.read_excel(filepath)
  df = _df.iloc[0: , [a+1, b+1, c+1, d+1]]
  return df


# 全ファイルで同じ形式で結合
  df = pd.DataFrame()
  for filepath in filepaths:
    _df = arrange(filepath)
    df = pd.concat([df, _df])
  df.to_excel('arranged_data.xlsx', index=False)


def mkws_byname(filepath):
  wb = openpyxl.load_workbook(filepath)
  ws = wb.worksheets[0]
  namelist = []
  for i in range(2, ws.max_row+1):
    name = ws.cell(i, 1).value
    if name not in namelist:
      namelist.append(name)
  for _ in namelist:
      wb.create_sheet(title=_)
  return wb


# 行ごとnamesheetに移動する
def get_row(filepath):
  wb = openpyxl.load_workbook(filepath)
  ws = wb.active

  row_no = 2
  for row_no in ws.max_row+1:
    name = ws.cell(row=row_no, column=1).value # 列1のセルから名前を取得
    for col in ws.iter_cols():
      for cell in col:
        ws[name].cell(row=max_row, column=1).value = cell.value # 同一bookのws[name]の最終行に追加
    row_no=+1
  return wb


# Paper size(n)selectボタンで設定
PAPERSIZE_LETTER = '1'
PAPERSIZE_LETTER_SMALL = '2'
PAPERSIZE_TABLOID = '3'
PAPERSIZE_LEDGER = '4'
PAPERSIZE_LEGAL = '5'
PAPERSIZE_STATEMENT = '6'
PAPERSIZE_EXECUTIVE = '7'
PAPERSIZE_A3 = '8'
PAPERSIZE_A4 = '9'
PAPERSIZE_A4_SMALL = '10'
PAPERSIZE_A5 = '11'


printinfo = PrintInfo()
# クラスからとってきた値を代入する関数にする, 例えば,papersize=10なら10を代入する
# nに用紙サイズ設定
def excel_print(filepath, printinfo): #　インスタンスごと渡す
  wb = openpyxl.load_workbook(filepath)
  for ws in wb.worksheets:
      ws.page_setup.orientation =  "f{printinfo.orientation}" #string
      ws.page_setup.paperSize =  printinfo.papersize #ws.PAPERSIZE_A4
      ws.page_setup.fitToWidth = printinfo.fitToWidth #0or1
      ws.page_setup.fitToHeight = printinfo.fitToHeight #0or1
      ws.sheet_properties.pageSetUpPr.fitToPage = True #True or False
      
  return wb

# 今日は,htmlとバックエンドのロジックの連関をやっていきます